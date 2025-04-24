import os
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from config import *
from utils import unlock_excel_file
import sys
import streamlit as st

def obtener_df():
    #directory_path = 'C:\\Users\\leonardo.mangold\\PycharmProjects\\promos_inteligencia_negocio\\ft_promos_automatico\\Cargar en Excel BI'
    directory_path = paths['files_para_BI']
    # List to store DataFrames
    dataframes = []

    # Iterate over all Excel files in the specified directory
    for file in os.listdir(directory_path):
        if file.endswith('.xlsx') or file.endswith('.xls'):
            file_path = os.path.join(directory_path, file)
            # Read each Excel file into a DataFrame
            df_aux = pd.read_excel(file_path, dtype={'ARTC_ARTC_COD':'str'})
            if 'Costo' not in df_aux.columns:
                df_aux['Costo'] = 0

            df_aux['ARTC_ARTC_COD'] = df_aux['ARTC_ARTC_COD'].astype(str)
            dataframes.append(df_aux)

    # Concatenate all DataFrames into one
    df = pd.concat(dataframes, ignore_index=True)
    df.rename({'ARTC_ARTC_COD':'Estadístico'}, axis=1, inplace=True)

    return df

def update_excel_with_dataframe(file_path, df):
    # Load the Excel workbook
    book = openpyxl.load_workbook(file_path)
    sheet_name = 'Ofertas'

    sheet = book[sheet_name]

    # Clear all data except the first row (the header)
    sheet.delete_rows(2, sheet.max_row - 1)

    # Write the DataFrame to the sheet, starting from row 2 (without headers)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)

    # Format column C as text
    for cell in sheet['C'][1:]:
        cell.number_format = '@'  # This sets the format to text (string)

    for col in ['A', 'B']:
        for cell in sheet[col][1:]:  # Skipping header
            if isinstance(cell.value, str):
                cell.value = pd.to_datetime(cell.value).date()
            cell.number_format = 'M/D/YYYY'

    # Save the workbook with the updates
    book.save(file_path)

def delete_all_files():

    #directory_path = 'C:\\Users\\leonardo.mangold\\PycharmProjects\\promos_inteligencia_negocio\\ft_promos_automatico\\Cargar en Excel BI'
    directory_path = paths['files_para_BI']
    for filename in os.listdir(directory_path):
        file_path = os.path.join(directory_path, filename)
        try:
            if os.path.isfile(file_path):
                os.remove(file_path)
        except Exception as e:
            st.write(f"Error deleting {file_path}: {e}")

def cargar_excel_BI(cursor, excel_path):

    st.write('')
    st.write('Inicia cargar_excel_BI')

    unlock_excel_file(excels['excel_BI'])  # Desbloqueo Excel Base BI

    excel_BI_df = pd.read_excel(
        excel_path,
        sheet_name='Ofertas',
        dtype = {'Estadístico':'str', 'Evento':'int'}
        )

    query = '''
    SELECT
        DISTINCT EVENTO_ID
    FROM
        MSTRDB.DWH.FT_PROMOS
    WHERE
        PROM_FECHA_FIN < CURRENT_DATE
    '''

    cursor.execute(query)
    promos_no_considerar = cursor.fetch_pandas_all()

    excel_BI_df = excel_BI_df[~excel_BI_df['Evento'].isin(promos_no_considerar['EVENTO_ID'])]

    excels_df = obtener_df()

    if len(excels_df.columns) == len(excel_BI_df.columns):
        excels_df.columns = excel_BI_df.columns
    else:
        st.write('Error: Distinto numero de columnas en Excel BI y excels a considerar')
        sys.exit()

    excel_BI_df = excel_BI_df[~excel_BI_df['Evento'].isin(excels_df['Evento'])]

    excel_BI_df = pd.concat([excels_df, excel_BI_df])

    excel_BI_df['Estadístico'] = excel_BI_df['Estadístico'].astype(str)
    excel_BI_df['Pronostico_Venta'] = 0
    excel_BI_df['Stock_inicial'] = 0
    excel_BI_df['Costo'] = 0

    for column in ['PVP OFERTA', 'Local_activo', 'Estiba_guia_operativa']:
        excel_BI_df[column].fillna(0, inplace = True)
        excel_BI_df[column] = excel_BI_df[column].astype(int)

    update_excel_with_dataframe(
        file_path=excel_path,
        df=excel_BI_df
    )

    delete_all_files()

    st.write('Termina cargar_excel_BI')
